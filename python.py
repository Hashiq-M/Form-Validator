import os
import openpyxl 
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

files = os.listdir('.')

excel_files = [file for file in files if file.endswith('.xlsx')]

if len(excel_files) < 2:
    raise ValueError("Not enough Excel files in the directory")

if len(excel_files) == 2:
    wb1 = openpyxl.load_workbook(excel_files[0])
    wb2 = openpyxl.load_workbook(excel_files[1])
    print(f"Loaded workbooks: {excel_files[0]} and {excel_files[1]}")

if len(excel_files) > 2:
    print("Multiple Excel files found. Please select two files to load:")
    for i, file in enumerate(excel_files):
        print(f"{i + 1}: {file}")
    
    first_choice = int(input("Enter the number for the first file: ")) - 1
    second_choice = int(input("Enter the number for the second file: ")) - 1
    
    wb1 = openpyxl.load_workbook(excel_files[first_choice])
    wb2 = openpyxl.load_workbook(excel_files[second_choice])
    print(f"Loaded workbooks: {excel_files[first_choice]} and {excel_files[second_choice]}")

wb_result = openpyxl.Workbook()
for wb in [wb1, wb2]:
    for sheet_name in wb.sheetnames:
        source = wb[sheet_name]
        target = wb_result.create_sheet(title=f"{sheet_name}")
        
        for row in source:
            for cell in row:
                target[cell.coordinate].value = cell.value

wb_result.remove(wb_result['Sheet'])
sh_result = wb_result.create_sheet(title="Comparison")
sh1 = wb1.active
sh2 = wb2.active
r1 = sh1.max_row
c1 = sh1.max_column
sh2_range = f"'{sh2.title}'!A1:{get_column_letter(sh2.max_column)}{sh2.max_row}"
for j in range(1, c1 + 1):
    sh_result.cell(1, 3*j-2, value=f"{sh1.title} - {sh1.cell(1, j).value}")
    sh_result.cell(1, 3*j-1, value=f"{sh2.title} - {sh2.cell(1, j).value}")
    sh_result.cell(1, 3*j, value="Comparison")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
for i in range(2, r1 + 1):
    for j in range(1, c1 + 1):
        col_letter1 = get_column_letter(3*j-2)
        col_letter2 = get_column_letter(3*j-1)
        col_letter3 = get_column_letter(3*j)
        
        sh_result.cell(i, 3*j-2, value=f"='{sh1.title}'!{get_column_letter(j)}{i}")
        
        vlookup_formula = f'=VLOOKUP(A{i},{sh2_range},{j},FALSE)'
        sh_result.cell(i, 3*j-1, value=vlookup_formula)
        
        compare_formula = f'=IF(OR(ISBLANK({col_letter1}{i}),ISBLANK({col_letter2}{i})),"NA",IF({col_letter1}{i}={col_letter2}{i},"True","False"))'
        comparison_cell = sh_result.cell(i, 3*j, value=compare_formula)
        
        comparison_cell.number_format = '@' 
        
        sh_result.conditional_formatting.add(f'{col_letter3}{i}',
            openpyxl.formatting.rule.CellIsRule(operator='equal', formula=['"True"'], stopIfTrue=True, fill=green_fill))
        sh_result.conditional_formatting.add(f'{col_letter3}{i}',
            openpyxl.formatting.rule.CellIsRule(operator='equal', formula=['"False"'], stopIfTrue=True, fill=red_fill))
        sh_result.conditional_formatting.add(f'{col_letter3}{i}',
            openpyxl.formatting.rule.CellIsRule(operator='equal', formula=['"NA"'], stopIfTrue=True, fill=yellow_fill))

result_filename = "Result.xlsx"
counter = 1
while os.path.exists(result_filename):
    result_filename = f"Result{counter}.xlsx"
    counter += 1

wb_result.save(result_filename)
print(f"{result_filename} completed")
